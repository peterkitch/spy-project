"""Run-once crunch rebuild orchestrator (thin shell; no auto-publish).

Chains the existing per-engine runners to rebuild a fixed set of broken K=6
secondary stacks while excluding a fixed set of broken tickers, verifying
exclusion at every artifact boundary, and STOPPING before any publication.

It is a thin shell over the existing CLIs -- it does NOT reimplement engine
logic and does NOT import the engines. Stages are invoked through an
injectable invoker boundary so tests can stub every stage:

  Stage 1  onepass_workbook_runner.py      (refresh allowed universe)
  Stage 2  impactsearch_workbook_runner.py (rebuild-secondary workbooks)
  Stage 3  stackbuilder_workbook_runner.py (re-select K=6 stacks)
  Stage 4  k6_recook.py --execute          (recook rebuilt stacks)

Modes:
  * Dry-run (default): Stage 0 preflight + plan only. No engines, no
    subprocess stage invocations, no network. Writes 00_preflight.json and
    run_plan.json and stops with status "dry_run_planned".
  * --execute (operator-run ONLY): runs Stages 1-4 with per-stage boundary
    checks, quarantine of the rebuild secondaries' prior canonical artifacts
    (move, never delete), and checkpoints. Fail-closed at every gate.

Stage-connection design: CANONICAL roots. The rebuilt stacks must REPLACE
the broken canonical output/stackbuilder/<SEC> stacks so the downstream
board picks them up; the prior broken artifacts are quarantined (moved) into
the run dir for rollback. The full-universe signal_library refresh in
Stage 1 is intended and operator-approved. Exclusion is guaranteed at the
selection boundaries (ImpactSearch primaries, StackBuilder candidacy, K6
member union), each proven by a boundary check that fails the run if any
excluded ticker is present.

After Stage 4 the orchestrator writes candidate artifacts to the canonical
roots / run dir but does NOT publish: no Blob, no promotion, no commit, no
push, no deploy. It records the exclusion + UNREBUILDABLE set for a later
operator manual master_tickers.txt edit (it never edits master_tickers.txt).

ASCII-only. Stdlib only (openpyxl used lazily only for the ImpactSearch
workbook boundary check during a real execute). Project-relative defaults.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable


SCHEMA_VERSION = "crunch_rebuild_run_v1"

DEFAULT_BLOCKED_FILE = "operator_inputs/crunch_blocked_tickers.txt"
DEFAULT_REBUILD_FILE = "operator_inputs/crunch_rebuild_secondaries.txt"
DEFAULT_RUN_BASE = "output/crunch_runs"
DEFAULT_STACKBUILDER_ROOT = "output/stackbuilder"
DEFAULT_IMPACTSEARCH_ROOT = "output/impactsearch"
DEFAULT_ONEPASS_ROOT = "output/onepass"
DEFAULT_K6_OUTPUT_ROOT = "output/k6_mtf"
DEFAULT_MASTER_TICKERS = "global_ticker_library/data/master_tickers.txt"
LOCK_NAME = ".crunch.lock"

# OnePass-reuse: default freshness window (hours) for reusing a prior crunch
# run's completed, validated OnePass evidence instead of re-running Stage 1.
# 168h (7 days) is a conservative weekly window and matches the pipeline's
# IMPACT_TRUST_MAX_AGE_HOURS staleness default (168h) used elsewhere for
# offline library reuse.
DEFAULT_REUSE_ONEPASS_MAX_AGE_HOURS = 168

SELECTED_BUILD = "selected_build.json"
SELECTED_BUILD_PINNED = "selected_build.pinned.json"
COMBO_FILENAME = "combo_k=6.json"

# Engine scripts this orchestrator launches; the process-conflict check MUST
# cover all of them (fail-closed if coverage is insufficient).
ENGINE_SCRIPTS = (
    "onepass.py", "onepass_workbook_runner.py",
    "impactsearch.py", "impactsearch_workbook_runner.py",
    "stackbuilder.py", "stackbuilder_workbook_runner.py",
    "k6_recook.py", "k6_mtf_history_producer.py", "k6_mtf_ranking_engine.py",
)

# A ticker token: optional leading caret, then symbol chars (letters, digits,
# dot, hyphen, caret, underscore, equals, colon) -- covers ^DJT, RPI-UN.TO,
# 011810.KS, DX-Y.NYB, BRK-A, BTC-USD.
_TICKER_RE = re.compile(r"^[\^A-Za-z0-9][A-Za-z0-9.\^_=:-]*$")
_PROTOCOL_RE = re.compile(r"\[[DI]\]$")


class CrunchError(Exception):
    """Fail-closed halt/refusal."""


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _parse_utc_timestamp(value: Any) -> datetime | None:
    """Parse an ISO-8601 UTC timestamp like '2026-06-06T19:10:16Z' (also
    tolerates an explicit +00:00 offset or a naive timestamp, which is assumed
    UTC). Returns a tz-aware UTC datetime, or None if unparseable."""
    s = str(value).strip()
    if not s:
        return None
    candidate = (s[:-1] + "+00:00") if s.endswith("Z") else s
    try:
        dt = datetime.fromisoformat(candidate)
    except ValueError:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _run_id(now: datetime) -> str:
    return now.strftime("%Y%m%dT%H%M%SZ")


def normalize_ticker(raw: Any) -> str:
    return str(raw).strip().upper()


def strip_protocol(token: str) -> str:
    return _PROTOCOL_RE.sub("", str(token).strip()).strip()


def _atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(prefix=path.name + ".", suffix=".part",
                              dir=str(path.parent))
    os.close(fd)
    tp = Path(tmp)
    try:
        tp.write_text(text, encoding="utf-8")
        os.replace(str(tp), str(path))
    finally:
        try:
            if tp.exists():
                tp.unlink()
        except OSError:
            pass


def _dump(obj: Any) -> str:
    return json.dumps(obj, indent=2, sort_keys=True) + "\n"


def _write_json(path: Path, obj: Any) -> None:
    _atomic_write_text(path, _dump(obj))


# ---------------------------------------------------------------------------
# Input loading + validation
# ---------------------------------------------------------------------------


def load_symbol_file(path: Path, *, label: str) -> list[str]:
    """Load a one-symbol-per-line file. Skips blank lines and '#' comments.
    Also accepts comma-separated tokens on a line. Validates each token is a
    recognizable symbol; raises CrunchError on any malformed token or an
    empty resulting list. Preserves first-seen order, de-duplicates."""
    if not path.is_file():
        raise CrunchError(f"{label} file not found: {path.as_posix()}")
    try:
        raw = path.read_text(encoding="utf-8")
    except OSError as exc:
        raise CrunchError(f"{label} file unreadable: {exc}") from exc
    if any(ord(ch) > 127 for ch in raw):
        raise CrunchError(f"{label} file contains non-ASCII bytes")
    seen: set[str] = set()
    out: list[str] = []
    for line in raw.splitlines():
        stripped = line.split("#", 1)[0].strip()
        if not stripped:
            continue
        for tok in stripped.split(","):
            t = tok.strip()
            if not t:
                continue
            if not _TICKER_RE.match(t):
                raise CrunchError(
                    f"{label} file has a malformed symbol: {t!r}")
            norm = normalize_ticker(t)
            if norm not in seen:
                seen.add(norm)
                out.append(norm)
    if not out:
        raise CrunchError(f"{label} file is empty after parsing")
    return out


def discover_current_secondaries(stackbuilder_root: Path) -> set[str]:
    """Current secondaries = subdirs of the stackbuilder root that carry a
    selected_build.json (matches how k6_recook discovers them). The
    '_progress' dir and any non-selected dir are ignored."""
    out: set[str] = set()
    if not stackbuilder_root.is_dir():
        return out
    for d in stackbuilder_root.iterdir():
        if d.is_dir() and (d / SELECTED_BUILD).is_file():
            out.add(normalize_ticker(d.name))
    return out


def load_master_universe(master_file: Path) -> list[str]:
    """Read the master ticker list (the OnePass default universe source).
    Raw-string parse (newline + comma), preserving literal NA / NAN."""
    if not master_file.is_file():
        return []
    raw = master_file.read_text(encoding="utf-8")
    out: list[str] = []
    seen: set[str] = set()
    for chunk in raw.replace("\n", ",").split(","):
        t = chunk.strip()
        if not t or t.startswith("#"):
            continue
        norm = normalize_ticker(t)
        if norm not in seen:
            seen.add(norm)
            out.append(norm)
    return out


# ---------------------------------------------------------------------------
# Lock
# ---------------------------------------------------------------------------


def _pid_alive(pid: int) -> bool | None:
    try:
        import psutil  # noqa: PLC0415
        return psutil.pid_exists(pid)
    except Exception:
        return None  # unknown -> caller treats conservatively


def acquire_lock(lock_path: Path, *, run_id: str, stage: str,
                 reclaim_stale: bool, now: datetime) -> None:
    payload = _dump({
        "pid": os.getpid(), "run_id": run_id, "stage": stage,
        "started_at": now.strftime("%Y-%m-%dT%H:%M:%SZ"),
    })
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        try:
            os.write(fd, payload.encode("utf-8"))
        finally:
            os.close(fd)
        return
    except FileExistsError:
        pass
    # Lock exists -- inspect holder.
    try:
        existing = json.loads(lock_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        existing = {}
    holder_pid = existing.get("pid")
    alive = _pid_alive(holder_pid) if isinstance(holder_pid, int) else None
    if alive is True:
        raise CrunchError(
            f"crunch lock held by live pid {holder_pid}; refusing")
    if alive is None:
        # Cannot prove dead -> conservative: treat as held unless reclaim.
        if not reclaim_stale:
            raise CrunchError(
                "crunch lock present and holder liveness unknown; rerun with "
                "--reclaim-stale-lock if it is stale")
    else:
        # Proven dead.
        if not reclaim_stale:
            raise CrunchError(
                f"crunch lock is stale (dead pid {holder_pid}); rerun with "
                "--reclaim-stale-lock to reclaim")
    # Reclaim.
    _atomic_write_text(lock_path, payload)


def release_lock(lock_path: Path) -> None:
    try:
        if lock_path.is_file():
            lock_path.unlink()
    except OSError:
        pass


# ---------------------------------------------------------------------------
# Process-conflict check (injectable)
# ---------------------------------------------------------------------------


def default_process_conflict_check(required_patterns: tuple) -> dict:
    """Read-only process enumeration mirroring the runners' approach. Returns
    {"status": "ok"|"blocked"|"insufficient", "conflicts": [...]}. "blocked"
    if a matching engine process is running; "insufficient" if processes
    cannot be enumerated (fail-closed for execute)."""
    own = os.getpid()
    cmdlines: list[str] = []
    try:
        import psutil  # noqa: PLC0415
        for proc in psutil.process_iter(["pid", "cmdline"]):
            if proc.info.get("pid") == own:
                continue
            cl = proc.info.get("cmdline") or []
            if cl:
                cmdlines.append(" ".join(str(x) for x in cl))
    except Exception:
        return {"status": "insufficient", "conflicts": [],
                "reason": "process enumeration unavailable"}
    low_patterns = [p.lower() for p in required_patterns]
    hits = []
    for cl in cmdlines:
        cll = cl.lower()
        for p in low_patterns:
            if p in cll:
                hits.append(p)
                break
    if hits:
        return {"status": "blocked", "conflicts": sorted(set(hits))}
    return {"status": "ok", "conflicts": []}


# ---------------------------------------------------------------------------
# Boundary checks: scan an artifact for any excluded ticker
# ---------------------------------------------------------------------------


_SEED_PROTOCOL_SUFFIX_RE = re.compile(r"-[DI]$")


def _member_tokens(token: str) -> set[str]:
    """Normalized ticker candidates from a scanned string. Handles plain
    symbols, bracket protocol forms (TICKER[D]/[I]), and embedded seed-name
    member forms (seedTC__AAA-D_DR8A.F-I), which join members with '_' and
    suffix each with -D/-I. Splits ONLY on '_' (no real ticker contains '_'),
    so it never substring-false-positives (e.g. FORD is not matched inside
    FORWARD; DX-Y.NYB is not split on its hyphens)."""
    out: set[str] = set()
    s = str(token).strip()
    if not s:
        return out
    # Whole-string variants (covers plain symbols and TICKER[D]/[I]).
    out.add(normalize_ticker(s))
    out.add(normalize_ticker(strip_protocol(s)))
    # Member-form variants: split on '_' and strip a trailing protocol.
    for piece in s.split("_"):
        p = piece.strip()
        if not p:
            continue
        out.add(normalize_ticker(p))
        no_bracket = strip_protocol(p)
        out.add(normalize_ticker(no_bracket))
        out.add(normalize_ticker(_SEED_PROTOCOL_SUFFIX_RE.sub("", no_bracket)))
    out.discard("")
    return out


def _collect_json_strings(obj: Any, out: list[str]) -> None:
    if isinstance(obj, str):
        out.append(obj)
    elif isinstance(obj, dict):
        for k, v in obj.items():
            out.append(str(k))
            _collect_json_strings(v, out)
    elif isinstance(obj, (list, tuple)):
        for v in obj:
            _collect_json_strings(v, out)


def scan_artifact_for_excluded(path: Path, excluded: set[str]) -> set[str]:
    """Return the set of excluded tickers found in an artifact. Handles
    .json / .txt (leaf/token scan) and .xlsx (every cell, via openpyxl).
    A missing file returns empty (callers decide if absence is itself an
    error)."""
    found: set[str] = set()
    if not path.is_file():
        return found
    suffix = path.suffix.lower()
    tokens: list[str] = []
    if suffix == ".json":
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            # Fall back to raw token scan.
            tokens = re.split(r"[^A-Za-z0-9.\^_=:\[\]-]+",
                              path.read_text(encoding="utf-8", errors="replace"))
        else:
            _collect_json_strings(data, tokens)
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook  # noqa: PLC0415
        except Exception as exc:
            raise CrunchError(
                f"cannot read xlsx boundary artifact (openpyxl missing): "
                f"{type(exc).__name__}") from exc
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if isinstance(cell, str):
                            tokens.append(cell)
        finally:
            wb.close()
    else:
        tokens = re.split(r"[^A-Za-z0-9.\^_=:\[\]-]+",
                          path.read_text(encoding="utf-8", errors="replace"))
    for tok in tokens:
        for variant in _member_tokens(tok):
            if variant in excluded:
                found.add(variant)
    return found


def assert_no_excluded_in_obj(obj: Any, excluded: set[str], *, stage: str) -> None:
    """Hard-STOP if any excluded ticker appears in an in-memory result object
    (e.g. a runner's per_ticker_results), using the same member-aware token
    normalization as the artifact scanner."""
    strings: list[str] = []
    _collect_json_strings(obj, strings)
    found: set[str] = set()
    for s in strings:
        for variant in _member_tokens(s):
            if variant in excluded:
                found.add(variant)
    if found:
        raise CrunchError(
            f"boundary check failed at {stage}: excluded tickers present in "
            f"runner result: {sorted(found)}")


def assert_no_excluded(paths: Iterable[Path], excluded: set[str],
                       *, stage: str) -> None:
    """Hard-STOP if any excluded ticker appears in any listed artifact."""
    offenders: dict[str, list[str]] = {}
    for p in paths:
        hits = scan_artifact_for_excluded(p, excluded)
        if hits:
            offenders[p.as_posix()] = sorted(hits)
    if offenders:
        raise CrunchError(
            f"boundary check failed at {stage}: excluded tickers present in "
            f"{json.dumps(offenders, sort_keys=True)}")


# ---------------------------------------------------------------------------
# Quarantine (move, never delete) -- rebuild secondaries only
# ---------------------------------------------------------------------------


def quarantine_paths(items: list[Path], dest_dir: Path) -> list[str]:
    """Move each existing path into dest_dir, preserving the basename.
    Returns the list of moved destinations (project-irrelevant absolute).
    Never deletes."""
    moved: list[str] = []
    if not items:
        return moved
    dest_dir.mkdir(parents=True, exist_ok=True)
    import shutil  # noqa: PLC0415
    for src in items:
        if not src.exists():
            continue
        target = dest_dir / src.name
        shutil.move(str(src), str(target))
        moved.append(target.as_posix())
    return moved


# ---------------------------------------------------------------------------
# Plan construction
# ---------------------------------------------------------------------------


def _csv(tickers: list[str]) -> str:
    return ",".join(tickers)


# Phase 6I-57 ImpactSearch optimized-runner env (non-secret config). All six
# are consumed by the engine/fast-path: IMPACT_REQUIRE_ZERO_PRIMARY_YF
# (impactsearch.py:644), IMPACT_INSTRUMENT_YF_CALLS (:722), IMPACT_TRUST_LIBRARY
# (:591), IMPACT_CALENDAR_GRACE_DAYS (:2983), IMPACT_MAX_WORKERS (:3724),
# IMPACT_TRUST_MAX_AGE_HOURS (signal_library/impact_fastpath.py:39). Injected
# ONLY into the ImpactSearch stage subprocess; other stages inherit os.environ
# unchanged (empty stage_env).
IMPACTSEARCH_STAGE_ENV = {
    "IMPACT_REQUIRE_ZERO_PRIMARY_YF": "1",
    "IMPACT_INSTRUMENT_YF_CALLS": "1",
    "IMPACT_TRUST_LIBRARY": "1",
    "IMPACT_TRUST_MAX_AGE_HOURS": "720",
    "IMPACT_CALENDAR_GRACE_DAYS": "30",
    "IMPACT_MAX_WORKERS": "8",
}

# Phase 6I-79 StackBuilder full K1-K12 build parity rationale (recorded in the
# plan for operator review).
STACKBUILDER_PARITY_RATIONALE = (
    "Full K1-K12 build parity with Phase 6I-79. The MTF site consumes the K6 "
    "slice downstream. --k-max 12 is full production-runner parity, not a "
    "narrowing."
)


def build_stage_commands(
    *, effective_rebuild: list[str], allowed_universe_file: Path,
    onepass_root: Path, impactsearch_root: Path, stackbuilder_root: Path,
    k6_output_root: Path, target_as_of: str | None,
    duration_budget_minutes: int | None, operator_budget_label: str | None,
    driver_run_id: str,
) -> list[dict]:
    """Resolve the exact argv for each stage from the verified runner
    contracts. Recorded in run_plan.json and used by the real invoker."""
    sec_csv = _csv(effective_rebuild)
    budget = ([] if duration_budget_minutes is None else
              ["--duration-budget-minutes", str(duration_budget_minutes)])
    label = ([] if not operator_budget_label else
             ["--operator-budget-label", operator_budget_label])
    target = [] if not target_as_of else ["--target-as-of", target_as_of]
    return [
        {
            "stage": "onepass",
            "script": "onepass_workbook_runner.py",
            "argv": [
                "--tickers-file", allowed_universe_file.as_posix(),
                "--output-dir", onepass_root.as_posix(),
                "--write", "--allow-network-fetch",
            ],
            "stage_env": {},
            "success_status": "ok",
        },
        {
            "stage": "impactsearch",
            "script": "impactsearch_workbook_runner.py",
            # ONE batch subprocess for all rebuild secondaries (the optimized
            # path relies on the warm in-process fast-path LRU across the
            # batch). Full-master primary universe unchanged; exclusion stays
            # input-withholding (the allowed-universe file = master minus the
            # exclusion set). Phase 6I-57 optimized flags: --use-multiprocessing
            # (parallel primary scoring) and --validation-mode legacy_fast
            # (skip durable validation).
            "argv": [
                "--secondaries", sec_csv,
                "--primary-source", "master_tickers_file",
                "--primary-tickers-file", allowed_universe_file.as_posix(),
                "--output-dir", impactsearch_root.as_posix(),
                "--use-multiprocessing",
                "--validation-mode", "legacy_fast",
                "--write", "--allow-network-fetch",
            ],
            "stage_env": dict(IMPACTSEARCH_STAGE_ENV),
            "success_status": "ok",
        },
        {
            "stage": "stackbuilder",
            "script": "stackbuilder_workbook_runner.py",
            # Phase 6I-79 full K1-K12 build parity.
            "argv": [
                "--secondaries", sec_csv,
                "--primary-source", "impact_xlsx",
                "--impact-xlsx-dir", impactsearch_root.as_posix(),
                "--outdir", stackbuilder_root.as_posix(),
                "--skip-durable-validation",
                "--jobs", "1",
                "--k-max", "12",
                "--exhaustive-k", "4",
                "--search", "beam",
                "--beam-width", "12",
                "--top-n", "20",
                "--bottom-n", "20",
                "--allow-decreasing",
                "--k-patience", "1",
                "--no-progress",
                "--write", "--allow-network-fetch", "--update-selected",
                *budget, *label,
            ],
            "stage_env": {},
            "parity_rationale": STACKBUILDER_PARITY_RATIONALE,
            "success_status": "ok",
        },
        {
            "stage": "k6_recook",
            "script": "k6_recook.py",
            "argv": [
                "--execute", "--allow-network-fetch",
                "--secondaries", sec_csv,
                "--driver-run-id", driver_run_id,
                "--stackbuilder-root", stackbuilder_root.as_posix(),
                "--output-root", k6_output_root.as_posix(),
                "--restage-all",
                *target,
            ],
            "stage_env": {},
            "success_status": "ok",
        },
    ]


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------


class CrunchOrchestrator:
    def __init__(
        self,
        *,
        project_root: Path,
        run_dir: Path,
        blocked_file: Path,
        rebuild_file: Path,
        stackbuilder_root: Path,
        impactsearch_root: Path,
        onepass_root: Path,
        k6_output_root: Path,
        master_tickers_file: Path,
        target_as_of: str | None,
        duration_budget_minutes: int | None,
        operator_budget_label: str | None,
        allow_network_fetch: bool,
        execute: bool,
        reclaim_stale_lock: bool,
        now: datetime,
        reuse_onepass_run_dir: Path | None = None,
        reuse_onepass_max_age_hours: int = DEFAULT_REUSE_ONEPASS_MAX_AGE_HOURS,
        publish_dry_run: bool = False,
        publish_fresh_ccc_records_file: Path | None = None,
        publish_prior_fixture: Path | None = None,
        publish_prior_promotion_manifest: Path | None = None,
        publish_prior_validation_sidecar: Path | None = None,
        publish_prior_ccc_verification_manifest: Path | None = None,
        publish: bool = False,
        operator_approved_publish: bool = False,
        invoker: Callable[[str, list[str], dict], dict] | None = None,
        conflict_check: Callable[[tuple], dict] | None = None,
        validator: Callable | None = None,
        joiner: Callable | None = None,
        combiner: Callable | None = None,
        stage9_runner: Callable | None = None,
    ) -> None:
        self.project_root = project_root
        self.run_dir = run_dir
        self.blocked_file = blocked_file
        self.rebuild_file = rebuild_file
        self.stackbuilder_root = stackbuilder_root
        self.impactsearch_root = impactsearch_root
        self.onepass_root = onepass_root
        self.k6_output_root = k6_output_root
        self.master_tickers_file = master_tickers_file
        self.target_as_of = target_as_of
        self.duration_budget_minutes = duration_budget_minutes
        self.operator_budget_label = operator_budget_label
        self.allow_network_fetch = allow_network_fetch
        self.execute = execute
        self.reclaim_stale_lock = reclaim_stale_lock
        self.now = now
        self.reuse_onepass_run_dir = reuse_onepass_run_dir
        self.reuse_onepass_max_age_hours = reuse_onepass_max_age_hours
        self._reuse: dict | None = None
        # Publish-dry-run tail (Stages 5-8). Publication boundary stays CLOSED.
        self.publish_dry_run = publish_dry_run
        self.publish_fresh_ccc_records_file = publish_fresh_ccc_records_file
        self.publish_prior_fixture = publish_prior_fixture
        self.publish_prior_promotion_manifest = publish_prior_promotion_manifest
        self.publish_prior_validation_sidecar = publish_prior_validation_sidecar
        self.publish_prior_ccc_verification_manifest = (
            publish_prior_ccc_verification_manifest)
        self._publish_gate: dict | None = None
        # Stage 9 operator-launched publish tail (real publication boundary;
        # operator-approved only). Default OFF keeps the closed-boundary modes
        # unchanged. The runner seam is stubbed in tests.
        self.publish = publish
        self.operator_approved_publish = operator_approved_publish
        self._stage9_runner = stage9_runner or self._default_stage9_runner
        self._stage9: dict | None = None
        self.invoker = invoker or self._default_invoker
        self.conflict_check = conflict_check or default_process_conflict_check
        # Injectable seams for the publish-dry-run tail (stubbed in tests; the
        # defaults are operator-only and never run during this dry-run task).
        self.validator = validator or self._default_validator
        self.joiner = joiner or self._default_joiner
        self.combiner = combiner or self._default_combiner
        self.run_id = run_dir.name
        self.lock_path = (run_dir.parent / LOCK_NAME)

    # --- real subprocess invoker (never used by tests) -------------------
    def _default_invoker(
        self, script: str, argv: list[str], stage_env: dict | None = None
    ) -> dict:
        cmd = [sys.executable, str(self.project_root / script), *argv]
        # Per-stage env injection: os.environ inherited verbatim, then the
        # stage's own keys layered on top. Empty/None stage_env => exact
        # os.environ (unchanged behavior for non-ImpactSearch stages).
        env = {**os.environ, **(stage_env or {})}
        proc = subprocess.run(cmd, capture_output=True, text=True, env=env)
        out = proc.stdout.strip()
        try:
            # The runners emit one JSON object on stdout; tolerate trailing
            # text by taking the last JSON object.
            start = out.rfind("\n{")
            blob = out[start + 1:] if start != -1 else out
            return json.loads(blob)
        except (json.JSONDecodeError, ValueError) as exc:
            raise CrunchError(
                f"stage {script} produced unparseable stdout "
                f"(exit {proc.returncode}): {type(exc).__name__}") from exc

    # --- envelope helpers -------------------------------------------------
    def _base_envelope(self, status: str, halted_at: str | None) -> dict:
        return {
            "schema_version": SCHEMA_VERSION,
            "run_id": self.run_id,
            "generated_at_utc": self.now.strftime("%Y-%m-%dT%H:%M:%SZ"),
            "mode": "execute" if self.execute else "dry_run",
            "status": status,
            "halted_at": halted_at,
            "publish_attempted": False,
            "blob_attempted": False,
            "promotion_attempted": False,
        }

    def _halt(self, halted_at: str, reason: str, preflight: dict | None) -> dict:
        env = self._base_envelope("halted", halted_at)
        env["reason"] = reason
        if preflight is not None:
            env["preflight"] = preflight
        _write_json(self.run_dir / "RUN_SUMMARY.json", env)
        return env

    # --- OnePass reuse (opt-in) -------------------------------------------
    def _rel_or_redact(self, p: Path) -> str:
        """Render a path project-relative when it lives under project_root,
        else redact to a non-absolute, non-leaking basename form. Never emits a
        local absolute path into proof output."""
        try:
            return p.resolve().relative_to(self.project_root.resolve()).as_posix()
        except (ValueError, OSError):
            return "<redacted>/" + p.name

    def _build_reuse_proof(self, allowed_universe: list[str],
                           excl_set: set[str]) -> dict:
        """Fail-closed validation of a prior crunch run's OnePass evidence.
        Returns a proof dict; proof['valid'] is True only if EVERY check
        passes. On the first failure it stops and records proof['reason'].
        Never raises; never mutates the prior run dir; all paths emitted are
        project-relative or redacted."""
        proof: dict = {
            "requested": True,
            "valid": False,
            "reason": None,
            "freshness_window_hours": self.reuse_onepass_max_age_hours,
            "no_onepass_subprocess": True,
        }
        prior = self.reuse_onepass_run_dir
        assert prior is not None  # only called when reuse requested

        # 1) prior dir exists.
        if not prior.is_dir():
            proof["source_run_dir"] = self._rel_or_redact(prior)
            proof["reason"] = (
                "prior reuse run dir not found: "
                + proof["source_run_dir"])
            return proof
        proof["source_run_dir"] = self._rel_or_redact(prior)

        # 2) prior dir is not the current run dir.
        try:
            is_same = prior.resolve() == self.run_dir.resolve()
        except OSError:
            is_same = False
        if is_same:
            proof["reason"] = "prior reuse run dir is the current run dir"
            return proof

        au_path = prior / "allowed_universe.txt"
        op_path = prior / "01_onepass.json"
        proof["source_allowed_universe"] = self._rel_or_redact(au_path)
        proof["source_onepass_json"] = self._rel_or_redact(op_path)

        # 3) prior allowed_universe.txt exists.
        if not au_path.is_file():
            proof["reason"] = "prior allowed_universe.txt is missing"
            return proof
        # 4) prior 01_onepass.json exists.
        if not op_path.is_file():
            proof["reason"] = "prior 01_onepass.json is missing"
            return proof

        try:
            data = json.loads(op_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            proof["reason"] = (
                "prior 01_onepass.json unreadable/unparseable: "
                + type(exc).__name__)
            return proof
        if not isinstance(data, dict):
            proof["reason"] = "prior 01_onepass.json is not a JSON object"
            return proof

        # 5) reused OnePass status is ok.
        status = data.get("status")
        proof["status"] = status
        if status != "ok":
            proof["reason"] = f"reused OnePass status is not ok: {status!r}"
            return proof

        # 6) reused OnePass has zero per-ticker errors (summary AND per-ticker).
        summary = data.get("summary")
        if not isinstance(summary, dict) or "error" not in summary:
            proof["reason"] = (
                "reused OnePass summary missing or has no 'error' count")
            return proof
        proof["summary"] = {k: summary.get(k) for k in ("error", "ok", "total")}
        if summary.get("error") != 0:
            proof["reason"] = (
                f"reused OnePass summary.error is not 0: "
                f"{summary.get('error')!r}")
            return proof
        per = data.get("per_ticker_results")
        if not isinstance(per, list) or not per:
            proof["reason"] = (
                "reused OnePass per_ticker_results missing or empty")
            return proof
        non_ok = [r.get("ticker") if isinstance(r, dict) else None
                  for r in per
                  if not isinstance(r, dict) or r.get("status") != "ok"]
        if non_ok:
            proof["reason"] = (
                f"reused OnePass has {len(non_ok)} per-ticker non-ok "
                f"result(s); first ticker: {non_ok[0]!r}")
            return proof

        # FIX 3: a read failure after the existence check fails closed instead
        # of leaking an uncaught OSError.
        try:
            au_raw = au_path.read_text(encoding="utf-8")
        except OSError as exc:
            proof["reason"] = (
                "prior allowed_universe.txt unreadable: " + type(exc).__name__)
            return proof
        prior_lines = [normalize_ticker(x) for x in au_raw.splitlines()
                       if x.strip()]

        # 7) no current blocked ticker in prior allowed_universe.txt, and none
        #    anywhere in the reused OnePass result (member-aware scan). Checked
        #    before the universe match so a blocked-ticker leak is reported as
        #    such (a blocked member would also fail the match, but the
        #    blocked-ticker reason is the more actionable refusal).
        in_au = sorted(t for t in prior_lines if t in excl_set)
        strings: list[str] = []
        _collect_json_strings(data, strings)
        found: set[str] = set()
        for s in strings:
            for variant in _member_tokens(s):
                if variant in excl_set:
                    found.add(variant)
        in_op = sorted(found)
        proof["blocked_ticker_scan"] = {
            "in_allowed_universe": in_au,
            "in_onepass_result": in_op,
        }
        if in_au or in_op:
            proof["blocked_scan_result"] = "blocked_ticker_present"
            proof["reason"] = (
                f"blocked ticker present (allowed_universe={in_au}, "
                f"onepass_result={in_op})")
            return proof
        proof["blocked_scan_result"] = "clean"

        # 8) prior allowed_universe.txt exactly matches the current recomputed
        #    allowed universe (normalized ordered lines).
        if prior_lines != allowed_universe:
            ps, cs = set(prior_lines), set(allowed_universe)
            proof["universe_match"] = False
            proof["universe_delta"] = {
                "prior_count": len(prior_lines),
                "current_count": len(allowed_universe),
                "in_prior_not_current": sorted(ps - cs)[:50],
                "in_current_not_prior": sorted(cs - ps)[:50],
                "same_set_order_differs": (ps == cs and prior_lines !=
                                           allowed_universe),
            }
            proof["reason"] = (
                "prior allowed_universe.txt does not match the current "
                "recomputed allowed universe")
            return proof
        proof["universe_match"] = True

        # 8b) FIX 2: the reused OnePass must prove FULL coverage of the current
        #     allowed universe -- not merely "zero errors". Summary counts and
        #     the ordered per-ticker list must each equal the current allowed
        #     universe exactly.
        n = len(allowed_universe)
        reused_tickers = [normalize_ticker(r.get("ticker")) for r in per]
        cov_problems: list[str] = []
        if summary.get("total") != n:
            cov_problems.append(f"summary.total={summary.get('total')!r} != {n}")
        if summary.get("ok") != n:
            cov_problems.append(f"summary.ok={summary.get('ok')!r} != {n}")
        if len(per) != n:
            cov_problems.append(f"len(per_ticker_results)={len(per)} != {n}")
        if reused_tickers != allowed_universe:
            cov_problems.append(
                "per_ticker_results tickers (ordered) do not equal the current "
                "allowed_universe")
        if cov_problems:
            rs, cs = set(reused_tickers), set(allowed_universe)
            proof["coverage_match"] = False
            proof["coverage_delta"] = {
                "expected_count": n,
                "summary_total": summary.get("total"),
                "summary_ok": summary.get("ok"),
                "per_ticker_count": len(per),
                "missing_from_result": sorted(cs - rs)[:50],
                "extra_in_result": sorted(rs - cs)[:50],
                "same_set_order_differs": (rs == cs and reused_tickers !=
                                           allowed_universe),
            }
            proof["reason"] = (
                "reused OnePass does not cover the current allowed universe: "
                + "; ".join(cov_problems))
            return proof
        proof["coverage_match"] = True

        # 9) freshness: prefer explicit completion timestamp, else mtime.
        parsed = _parse_utc_timestamp(data.get("end_timestamp_utc"))
        ts_source = "end_timestamp_utc" if parsed is not None else None
        if parsed is None:
            try:
                parsed = datetime.fromtimestamp(op_path.stat().st_mtime,
                                                tz=timezone.utc)
                ts_source = "file_mtime"
            except OSError:
                parsed = None
        if parsed is None:
            proof["reason"] = (
                "reused OnePass has no parseable completion timestamp and "
                "file mtime is unavailable")
            return proof
        proof["timestamp_source"] = ts_source
        proof["parsed_timestamp_utc"] = parsed.strftime("%Y-%m-%dT%H:%M:%SZ")
        age_hours = (self.now - parsed).total_seconds() / 3600.0
        proof["age_hours"] = round(age_hours, 3)
        if age_hours < 0:
            proof["reason"] = (
                f"reused OnePass timestamp is in the future "
                f"(age {age_hours:.3f}h, source {ts_source})")
            return proof
        if age_hours > self.reuse_onepass_max_age_hours:
            proof["fresh"] = False
            proof["reason"] = (
                f"reused OnePass is stale: age {age_hours:.2f}h > window "
                f"{self.reuse_onepass_max_age_hours}h (timestamp source "
                f"{ts_source})")
            return proof
        proof["fresh"] = True

        # All checks passed.
        proof["valid"] = True
        proof["summary_text"] = (
            f"Stage 1 reused from prior run {proof['source_run_dir']}: "
            f"status=ok, {summary.get('ok')}/{summary.get('total')} tickers "
            f"ok, allowed-universe match, blocked-ticker scan clean, age "
            f"{proof['age_hours']}h <= {self.reuse_onepass_max_age_hours}h "
            f"(timestamp source: {ts_source}). No OnePass subprocess will be "
            f"invoked.")
        return proof

    # --- Stage 0 ----------------------------------------------------------
    def preflight(self) -> dict:
        self.run_dir.mkdir(parents=True, exist_ok=True)
        blocked = load_symbol_file(self.blocked_file, label="blocked-tickers")
        rebuild = load_symbol_file(self.rebuild_file,
                                   label="rebuild-secondaries")
        exclusion = sorted(set(blocked))
        excl_set = set(exclusion)
        current_secs = discover_current_secondaries(self.stackbuilder_root)
        # UNREBUILDABLE = exclusion tickers that are current secondaries.
        unrebuildable = sorted(s for s in exclusion if s in current_secs)
        requested = sorted(set(rebuild))
        effective = sorted(s for s in requested if s not in excl_set)
        master = load_master_universe(self.master_tickers_file)
        allowed_universe = [t for t in master if t not in excl_set]

        # Write the run-scoped allowed-universe file (never master_tickers).
        allowed_file = self.run_dir / "allowed_universe.txt"
        _atomic_write_text(allowed_file,
                           "".join(t + "\n" for t in allowed_universe))

        # OnePass reuse (opt-in): validate prior evidence against the CURRENT
        # recomputed universe/exclusion. Never bypasses the recomputation above.
        reuse = None
        if self.reuse_onepass_run_dir is not None:
            reuse = self._build_reuse_proof(allowed_universe, excl_set)

        commands = build_stage_commands(
            effective_rebuild=effective,
            allowed_universe_file=allowed_file,
            onepass_root=self.onepass_root,
            impactsearch_root=self.impactsearch_root,
            stackbuilder_root=self.stackbuilder_root,
            k6_output_root=self.k6_output_root,
            target_as_of=self.target_as_of,
            duration_budget_minutes=self.duration_budget_minutes,
            operator_budget_label=self.operator_budget_label,
            driver_run_id=self.run_id,
        )

        gates = {
            "execute": self.execute,
            "allow_network_fetch": self.allow_network_fetch,
            "duration_budget_minutes": self.duration_budget_minutes,
            "operator_budget_label": self.operator_budget_label,
            "target_as_of": self.target_as_of,
        }
        missing_gates = []
        if self.execute:
            if not self.allow_network_fetch:
                missing_gates.append("--allow-network-fetch")
            if not self.duration_budget_minutes:
                missing_gates.append("--duration-budget-minutes")
            if not self.operator_budget_label:
                missing_gates.append("--operator-budget-label")
            if not self.target_as_of:
                missing_gates.append("--target-as-of")

        conflict = self.conflict_check(ENGINE_SCRIPTS)

        warnings = []
        if not self.target_as_of:
            warnings.append("target-as-of not set (required for --execute)")
        if not allowed_universe:
            warnings.append(
                "allowed universe is empty (master_tickers not found?)")
        if not effective:
            warnings.append("effective rebuild set is empty")

        preflight = {
            "schema_version": SCHEMA_VERSION,
            "run_id": self.run_id,
            "exclusion_set": exclusion,
            "exclusion_count": len(exclusion),
            "requested_rebuild_set": requested,
            "effective_rebuild_set": effective,
            "unrebuildable_set": unrebuildable,
            "allowed_universe_size": len(allowed_universe),
            "allowed_universe_file": allowed_file.as_posix(),
            "current_secondary_count": len(current_secs),
            "stage_commands": commands,
            "execution_gates": gates,
            "missing_execute_gates": missing_gates,
            "process_conflict": conflict,
            "stage_connection": {
                "design": "canonical_roots_with_quarantine",
                "rationale": (
                    "rebuilt stacks must replace broken canonical "
                    "output/stackbuilder/<SEC>; prior artifacts are moved to "
                    "the run dir quarantine (never deleted). Stage 1 "
                    "full-universe signal_library refresh is operator-approved."
                ),
                "canonical_mutation_under_execute": [
                    "signal_library/data/stable (Stage 1, full allowed universe)",
                    "output/impactsearch/<rebuild SEC> (Stage 2)",
                    "output/stackbuilder/<rebuild SEC> (Stage 3)",
                    "cache/results, price_cache/daily, signal_library "
                    "(Stage 4 k6_recook A/Aprime/B)",
                    "output/k6_mtf/<run> (Stage 4, run-scoped)",
                ],
            },
            "runner_contract_notes": [
                "No engine honors a ban/exclude flag; exclusion is by "
                "input-withholding (allowed-universe file = master minus "
                "exclusion; ImpactSearch primaries from that file; StackBuilder "
                "primaries from the freshly rebuilt impact xlsx).",
                "ImpactSearch keep-last carry-forward defeated by quarantining "
                "each rebuild secondary's prior workbook before Stage 2.",
                "k6_recook --restage-all prevents a stale selected_build.json "
                "from carrying a broken member through Stage 0.",
                "k6_recook run WITHOUT --allow-stage-a-exclusions so any Stage "
                "A unavailability is a hard STOP (no silent member shrink).",
            ],
            "warnings": warnings,
        }
        run_plan = {
            "schema_version": SCHEMA_VERSION,
            "run_id": self.run_id,
            "mode": "execute" if self.execute else "dry_run",
            "stage_commands": commands,
            "exclusion_set": exclusion,
            "effective_rebuild_set": effective,
            "unrebuildable_set": unrebuildable,
            "execution_gates": gates,
        }
        # Reuse metadata only when reuse is requested (otherwise the preflight
        # and run_plan shapes are exactly as today).
        if reuse is not None:
            stage1_view = {
                "stage": "onepass",
                "action": "reused" if reuse.get("valid") else "refused",
                "onepass_subprocess_invoked": False,
                "proof": reuse,
            }
            preflight["onepass_reuse"] = reuse
            run_plan["onepass_reuse"] = reuse
            run_plan["stage1_onepass"] = stage1_view
        _write_json(self.run_dir / "00_preflight.json", preflight)
        _write_json(self.run_dir / "run_plan.json", run_plan)
        # Stash for stages.
        self._excl_set = excl_set
        self._exclusion = exclusion
        self._requested = requested
        self._effective = effective
        self._unrebuildable = unrebuildable
        self._commands = commands
        self._conflict = conflict
        self._missing_gates = missing_gates
        self._reuse = reuse
        return preflight

    def _write_manual_edit_outputs(self) -> None:
        manual = sorted(set(self._exclusion) | set(self._unrebuildable))
        _atomic_write_text(
            self.run_dir / "broken_tickers_for_manual_master_ticker_edit.txt",
            "".join(t + "\n" for t in manual))
        _write_json(
            self.run_dir / "broken_tickers_for_manual_master_ticker_edit.json",
            {
                "schema_version": SCHEMA_VERSION,
                "run_id": self.run_id,
                "note": ("operator manually edits master_tickers.txt; the "
                         "orchestrator never edits it"),
                "exclusion_set": sorted(self._exclusion),
                "unrebuildable_secondaries": sorted(self._unrebuildable),
                "manual_master_ticker_removal_candidates": manual,
            })

    def _same_as_run_dir(self, other: Path) -> bool:
        try:
            return other.resolve() == self.run_dir.resolve()
        except OSError:
            return False

    # --- top-level run ----------------------------------------------------
    def run(self) -> dict:
        # FIX 1: refuse a reuse source dir equal to the current run dir BEFORE
        # touching it. preflight() would otherwise mkdir the run dir and write
        # allowed_universe.txt / 00_preflight.json / run_plan.json into it --
        # mutating the prior evidence dir before the refusal fires. Return a
        # halted envelope directly (NOT via _halt, which would write
        # RUN_SUMMARY.json into that same dir). No mkdir, no lock, no invoker.
        if self.reuse_onepass_run_dir is not None and self._same_as_run_dir(
                self.reuse_onepass_run_dir):
            env = self._base_envelope("halted", "reuse_onepass")
            env["reason"] = (
                "OnePass reuse refused: the reuse source dir cannot equal the "
                "current run dir (refused before any run-dir write)")
            return env

        # Stage 0 (both modes).
        try:
            preflight = self.preflight()
        except CrunchError as exc:
            self.run_dir.mkdir(parents=True, exist_ok=True)
            return self._halt("preflight", str(exc), None)

        # Lock (both modes); released at the end / on halt.
        try:
            acquire_lock(self.lock_path, run_id=self.run_id, stage="stage0",
                         reclaim_stale=self.reclaim_stale_lock, now=self.now)
        except CrunchError as exc:
            return self._halt("lock", str(exc), preflight)

        try:
            # Conflict gate: blocked -> STOP (both modes); insufficient ->
            # STOP for execute, warning for dry-run.
            cstat = self._conflict.get("status")
            if cstat == "blocked":
                return self._halt("process_conflict",
                                  "conflicting engine process detected",
                                  preflight)
            if cstat == "insufficient" and self.execute:
                return self._halt(
                    "process_conflict",
                    "process-conflict coverage insufficient (cannot enumerate "
                    "processes); refusing execute", preflight)

            # OnePass-reuse gate (both modes): if reuse was requested it must
            # validate, or the whole run STOPs fail-closed.
            if self.reuse_onepass_run_dir is not None and not (
                    self._reuse and self._reuse.get("valid")):
                reason = ((self._reuse or {}).get("reason")
                          or "reuse validation failed")
                return self._halt("reuse_onepass",
                                  "OnePass reuse refused: " + reason, preflight)

            if not self.execute:
                env = self._base_envelope("dry_run_planned", None)
                env["preflight"] = preflight
                env["note"] = ("dry-run: no stages invoked, no network, no "
                               "engine runs")
                self._write_manual_edit_outputs()
                _write_json(self.run_dir / "RUN_SUMMARY.json", env)
                return env

            # ----- execute path (operator-run only) -----
            if self._missing_gates:
                return self._halt(
                    "execute_gates",
                    "missing required execute gates: "
                    + ", ".join(self._missing_gates), preflight)
            # Stage 9: refuse early (before Stage 1 compute) if publish is
            # requested without approval or if the Stage 9 preflight fails.
            if self.publish:
                if not self.operator_approved_publish:
                    return self._halt(
                        "publish_gates",
                        "--publish requires --operator-approved-publish",
                        preflight)
                try:
                    self._stage9_preflight()
                except Exception as exc:  # noqa: BLE001 - any preflight failure
                    return self._halt("stage9_preflight", str(exc), preflight)
            return self._run_execute(preflight)
        finally:
            release_lock(self.lock_path)

    # --- execute stages ---------------------------------------------------
    def _stage_cmd(self, stage: str) -> dict:
        for c in self._commands:
            if c["stage"] == stage:
                return c
        raise CrunchError(f"no command for stage {stage}")

    def _require_ok(self, stage: str, result: dict) -> None:
        status = result.get("status")
        if status != "ok":
            raise CrunchError(
                f"stage {stage} non-ok status: {status!r} "
                f"(halted_at={result.get('halted_at')})")

    def _assert_impactsearch_parity(self, result: dict) -> None:
        """Fail closed unless the ImpactSearch run matched the optimized-runner
        profile: validation_mode == 'legacy_fast', durable_validation_ran is
        False, and primary_yfinance_fetch_count == 0. The runner reports these
        three fields PER SECONDARY (no top-level aggregate); enforce every
        per-secondary entry and also enforce any aggregate field that happens
        to be present. A missing required field is a hard STOP -- we never pass
        a run we cannot prove matched the profile."""

        def _check(scope: str, holder: dict) -> None:
            mode = holder.get("validation_mode")
            if mode != "legacy_fast":
                raise CrunchError(
                    f"impactsearch parity ({scope}): validation_mode "
                    f"{mode!r} != 'legacy_fast'")
            ran = holder.get("durable_validation_ran")
            if ran is not False:
                raise CrunchError(
                    f"impactsearch parity ({scope}): durable_validation_ran "
                    f"{ran!r} != False")
            fetched = holder.get("primary_yfinance_fetch_count")
            if fetched != 0:
                raise CrunchError(
                    f"impactsearch parity ({scope}): "
                    f"primary_yfinance_fetch_count {fetched!r} != 0")

        # Aggregate fields, enforced only where present (the runner currently
        # exposes none at the top level; this future-proofs if it adds them).
        agg_keys = ("validation_mode", "durable_validation_ran",
                    "primary_yfinance_fetch_count")
        if any(k in result for k in agg_keys):
            for k in agg_keys:
                if k not in result:
                    raise CrunchError(
                        f"impactsearch parity (aggregate): partial aggregate "
                        f"fields present but {k!r} missing")
            _check("aggregate", result)

        per = result.get("per_ticker_results")
        if not isinstance(per, list) or not per:
            raise CrunchError(
                "impactsearch parity: executed result is missing a non-empty "
                "per_ticker_results list")
        required = ("validation_mode", "durable_validation_ran",
                    "primary_yfinance_fetch_count")
        for idx, rec in enumerate(per):
            if not isinstance(rec, dict):
                raise CrunchError(
                    f"impactsearch parity: per_ticker_results[{idx}] is not "
                    "an object")
            sec = rec.get("secondary", f"#{idx}")
            for k in required:
                if k not in rec:
                    raise CrunchError(
                        f"impactsearch parity (secondary {sec!r}): required "
                        f"field {k!r} missing")
            _check(f"secondary {sec!r}", rec)

    def _assert_stackbuilder_parity(self, result: dict) -> None:
        """Fail closed unless the executed StackBuilder result envelope proves
        the intended Phase 6I-79 optimized build shape. Result-envelope ONLY:
        the top-level effective_config records the build profile, summary
        records the run outcome, and per_secondary_results records each
        secondary's status. A missing/wrong field is a hard STOP -- we never
        accept an executed StackBuilder run we cannot prove ran the optimized
        shape. This guards executed build shape; it does NOT read per-secondary
        run_manifest.json and is not a publish-time durable-validation check."""
        cfg = result.get("effective_config")
        if not isinstance(cfg, dict):
            raise CrunchError(
                "stackbuilder parity: effective_config missing or not a "
                f"mapping (got {type(cfg).__name__})")
        # (key, expected). Booleans are matched by identity so 1/0 cannot
        # masquerade as True/False; ints/strs by equality.
        expected = (
            ("skip_durable_validation", True),
            ("k_max", 12),
            ("exhaustive_k", 4),
            ("search", "beam"),
            ("beam_width", 12),
            ("top_n", 20),
            ("bottom_n", 20),
            ("k_patience", 1),
            ("allow_decreasing", True),
            ("jobs", 1),
        )
        for key, want in expected:
            if key not in cfg:
                raise CrunchError(
                    f"stackbuilder parity: effective_config missing {key!r}")
            got = cfg.get(key)
            if isinstance(want, bool):
                if got is not want:
                    raise CrunchError(
                        f"stackbuilder parity: effective_config {key}={got!r} "
                        f"is not {want!r}")
            elif got != want:
                raise CrunchError(
                    f"stackbuilder parity: effective_config {key}={got!r} "
                    f"!= {want!r}")

        summary = result.get("summary")
        if not isinstance(summary, dict):
            raise CrunchError(
                "stackbuilder parity: summary missing or not a mapping")
        if "error" not in summary:
            raise CrunchError("stackbuilder parity: summary.error missing")
        if summary.get("error") != 0:
            raise CrunchError(
                f"stackbuilder parity: summary.error="
                f"{summary.get('error')!r} != 0")

        per = result.get("per_secondary_results")
        if not isinstance(per, list) or not per:
            raise CrunchError(
                "stackbuilder parity: per_secondary_results missing or not a "
                "non-empty list")
        for idx, rec in enumerate(per):
            if not isinstance(rec, dict):
                raise CrunchError(
                    f"stackbuilder parity: per_secondary_results[{idx}] is "
                    "not an object")
            sec = rec.get("secondary", f"#{idx}")
            if "status" not in rec:
                raise CrunchError(
                    f"stackbuilder parity (secondary {sec!r}): status missing")
            if rec.get("status") != "ok":
                raise CrunchError(
                    f"stackbuilder parity (secondary {sec!r}): status "
                    f"{rec.get('status')!r} != 'ok'")

    def _run_execute(self, preflight: dict) -> dict:
        excl = self._excl_set
        rebuild = self._effective
        checkpoints: dict[str, Any] = {}
        try:
            # Stage 1 -- OnePass (allowed universe; exclusion omitted), OR
            # reuse of a prior run's already-validated OnePass evidence.
            if self.reuse_onepass_run_dir is not None and self._reuse \
                    and self._reuse.get("valid"):
                # Reuse proof was validated in preflight (status ok, zero
                # per-ticker errors, exact universe match, blocked-ticker scan
                # clean, fresh). Write a reuse checkpoint + a reused
                # 01_onepass.json in the NEW run dir; touch nothing in the
                # prior run dir. No OnePass subprocess is invoked.
                _write_json(self.run_dir / "00_onepass_reuse_proof.json",
                            self._reuse)
                r1 = {
                    "status": "ok",
                    "mode": "reused",
                    "source_run_dir": self._reuse.get("source_run_dir"),
                    "reused_onepass_json": self._reuse.get(
                        "source_onepass_json"),
                    "reused_allowed_universe": self._reuse.get(
                        "source_allowed_universe"),
                    "proof": self._reuse,
                }
                _write_json(self.run_dir / "01_onepass.json", r1)
                checkpoints["onepass"] = "reused"
            else:
                c1 = self._stage_cmd("onepass")
                r1 = self.invoker("onepass_workbook_runner.py",
                                  c1["argv"], c1.get("stage_env") or {})
                self._require_ok("onepass", r1)
                # Boundary: the runner result itself (per_ticker_results) must
                # carry no excluded ticker, even if the canonical manifest is
                # clean; plus the result-reported workbook/manifest files and
                # the canonical manifest.
                assert_no_excluded_in_obj(
                    r1.get("per_ticker_results"), excl, stage="onepass")
                op_paths = [self.onepass_root / "onepass.xlsx.manifest.json"]
                for key in ("workbook_path", "manifest_path"):
                    val = r1.get(key)
                    if val:
                        op_paths.append(self._abspath(val))
                assert_no_excluded(op_paths, excl, stage="onepass")
                _write_json(self.run_dir / "01_onepass.json", r1)
                checkpoints["onepass"] = r1.get("status")

            # Stage 2 -- ImpactSearch (quarantine first, primaries exclude set).
            q_imp = self._quarantine_impactsearch(rebuild)
            c2 = self._stage_cmd("impactsearch")
            r2 = self.invoker("impactsearch_workbook_runner.py",
                              c2["argv"], c2.get("stage_env") or {})
            self._require_ok("impactsearch", r2)
            # Optimized-runner parity boundary: legacy_fast, no durable
            # validation, zero primary yfinance fetches (per secondary).
            self._assert_impactsearch_parity(r2)
            assert_no_excluded(self._impactsearch_artifacts(rebuild),
                               excl, stage="impactsearch")
            _write_json(self.run_dir / "02_impactsearch.json",
                        {"result": r2, "quarantined": q_imp})
            checkpoints["impactsearch"] = r2.get("status")

            # Stage 3 -- StackBuilder (quarantine first, re-select members).
            self._assert_no_blocking_pins(rebuild)
            q_sb = self._quarantine_stackbuilder(rebuild)
            c3 = self._stage_cmd("stackbuilder")
            r3 = self.invoker("stackbuilder_workbook_runner.py",
                              c3["argv"], c3.get("stage_env") or {})
            self._require_ok("stackbuilder", r3)
            # Optimized-runner parity boundary: prove the executed result
            # envelope confirms the Phase 6I-79 build shape (K1-K12 beam,
            # durable validation skipped, all secondaries ok).
            self._assert_stackbuilder_parity(r3)
            assert_no_excluded(self._stackbuilder_artifacts(rebuild),
                               excl, stage="stackbuilder")
            _write_json(self.run_dir / "03_stackbuilder.json",
                        {"result": r3, "quarantined": q_sb})
            checkpoints["stackbuilder"] = r3.get("status")

            # Stage 4 -- k6_recook (restage; Stage A authoritative).
            c4 = self._stage_cmd("k6_recook")
            r4 = self.invoker("k6_recook.py",
                              c4["argv"], c4.get("stage_env") or {})
            self._require_ok("k6_recook", r4)
            stagea = r4.get("stageA") or {}
            if stagea.get("excluded_secondaries"):
                raise CrunchError(
                    "k6_recook Stage A excluded secondaries; refusing silent "
                    "member shrink: "
                    + json.dumps(stagea.get("excluded_secondaries"),
                                 sort_keys=True))
            assert_no_excluded(self._k6_artifacts(r4), excl, stage="k6_recook")
            _write_json(self.run_dir / "04_k6_recook.json", r4)
            checkpoints["k6_recook"] = r4.get("status")

            # Stages 5-8 -- gated publish-dry-run tail. The publication boundary
            # stays CLOSED: no Blob upload/GET, no promote CLI/--write, no
            # public fixture write, no commit/push/deploy. Runs ONLY with
            # --publish-dry-run; any failure routes through the except below.
            if self.publish:
                # Stage 9 -- operator-launched publish tail (real publication).
                self._stage9 = self._run_stage9_publish_tail(r4, rebuild)
                checkpoints["stage9_publish"] = self._stage9.get("status")
                if self._stage9.get("status") == "refused":
                    raise CrunchError(
                        "stage 9 publish refused at "
                        f"{self._stage9.get('stage')}: "
                        f"{self._stage9.get('reason')}")
            elif self.publish_dry_run:
                self._publish_gate = self._run_publish_dry_run_tail(r4, rebuild)
                checkpoints["publish_dry_run"] = "ok"
        except CrunchError as exc:
            self._write_manual_edit_outputs()
            env = self._halt("execute", str(exc), preflight)
            env["checkpoints"] = checkpoints
            _write_json(self.run_dir / "RUN_SUMMARY.json", env)
            return env

        self._write_manual_edit_outputs()
        if self.publish:
            status = "completed_publish"
        elif self.publish_dry_run:
            status = "completed_publish_dry_run"
        else:
            status = "completed_no_publish"
        env = self._base_envelope(status, None)
        env["preflight"] = preflight
        env["checkpoints"] = checkpoints
        env["candidate_artifacts_root"] = self.k6_output_root.as_posix()
        if self.publish_dry_run and self._publish_gate is not None:
            env["publish_dry_run"] = self._publish_gate
        if self.publish and self._stage9 is not None:
            env["stage9_publish"] = self._stage9
        _write_json(self.run_dir / "RUN_SUMMARY.json", env)
        return env

    # --- publish-dry-run tail (Stages 5-8; publication boundary CLOSED) ----
    def _rel(self, p: Path) -> str:
        try:
            return p.resolve().relative_to(self.project_root.resolve()).as_posix()
        except (ValueError, OSError):
            return "<redacted>/" + p.name

    def _prior_fixture_path(self) -> Path:
        return (self.publish_prior_fixture if self.publish_prior_fixture is not None
                else self.project_root / "frontend" / "public" / "fixtures"
                / "k6_mtf_ranking.json")

    def _prior_promotion_manifest_path(self) -> Path:
        return (self.publish_prior_promotion_manifest
                if self.publish_prior_promotion_manifest is not None
                else self.project_root / "frontend" / "public" / "fixtures"
                / "k6_mtf_ranking.promotion_manifest.json")

    def _assert_validation_sidecar(self, sidecar: Any,
                                   effective: list[str]) -> None:
        """Fail closed unless the Stage-5 sidecar covers EXACTLY the built
        secondaries with strict k6_mtf:<SEC> strategy ids and binds to this
        run id. Delegates to the shared module-level assertion."""
        _assert_validation_sidecar_covers(sidecar, effective, run_id=self.run_id)

    def _publish_gate_report(self, *, status: str, reason: str | None,
                             candidate: dict | None,
                             join_meta: dict | None) -> dict:
        counts = None
        if isinstance(candidate, dict):
            counts = {k: candidate.get(k) for k in (
                "merged_row_count", "board_validated_count", "not_validated_count",
                "carried_count", "fresh_count", "net_new_count",
                "stage_a_excluded_count", "ccc_record_count")}
        return {
            "stage": "publish_gate",
            "mode": "publish_dry_run",
            "status": status,
            "reason": reason,
            "join": join_meta,
            "candidate_artifacts": (candidate.get("paths")
                                    if isinstance(candidate, dict) else None),
            "candidate_counts": counts,
            "would_be_publish_plan": [
                "build slim fixture + CCC manifest (NOT written to "
                "frontend/public)",
                "upload + GET-verify CCC sidecars for built rows (NOT executed)",
                "promote_k6_mtf_artifact.py --public --write --operator-approved "
                "(NOT executed)",
                "git commit + push of the public fixture (NOT executed)",
            ],
            # Explicit closed-boundary flags (this stage performs none of these).
            "no_blob_upload": True,
            "no_blob_get": True,
            "no_promote_cli_invoked": True,
            "no_promote_write": True,
            "no_operator_approved": True,
            "no_public_fixture_write": True,
            "no_commit": True,
            "no_push": True,
            "no_deploy": True,
        }

    def _load_fresh_ccc_records(self, path: Path) -> list:
        if not Path(path).is_file():
            raise CrunchError(
                f"fresh CCC records file not found: {self._rel(Path(path))}")
        try:
            data = json.loads(Path(path).read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise CrunchError(
                f"fresh CCC records file unreadable/invalid JSON: "
                f"{type(exc).__name__}") from exc
        records = data.get("records") if isinstance(data, dict) else data
        if not isinstance(records, list) or not records:
            raise CrunchError(
                "fresh CCC records file must be a non-empty list (or an object "
                "with a non-empty 'records' list)")
        return records

    # --- fresh-row path normalization (Stage 6 -> Stage 7) ----------------
    # Single source of truth lives at module scope (_RERANK_FRESH_ROW_PATH_FIELDS
    # / _normalize_one_output_path / _normalize_fresh_row_paths); these methods
    # delegate so the dry-run tail and the re-rank publish seam share behavior.
    def _normalize_one_output_path(self, sec: str, field: str, value: Any) -> str:
        """Delegate to the shared module-level normalizer (bound to
        self.project_root)."""
        return _normalize_one_output_path(
            sec, field, value, project_root=self.project_root)

    def _normalize_publish_fresh_row_paths(self, rows: list) -> tuple:
        """Delegate to the shared module-level fresh-row normalizer (bound to
        self.project_root)."""
        return _normalize_fresh_row_paths(rows, project_root=self.project_root)

    def _run_publish_dry_run_tail(self, r4: dict, rebuild: list[str]) -> dict:
        """Stages 5-8: validation -> join -> combine/proof -> publish-gate
        dry-run. Writes numbered checkpoints under the run dir and a would-be
        publish plan, then returns the gate report. Performs NO network, Blob,
        promote, fixture write, commit, push, or deploy. Any failure raises
        CrunchError (routed to the existing halt path)."""
        from crunch_combine_proof import CombineError  # noqa: PLC0415
        from utils.react_publish.k6_mtf_validation_join import (  # noqa: PLC0415
            ValidationJoinError as _ValidationJoinError)
        effective = list(rebuild)

        # ----- Stage 5: validation (injectable seam) -----
        sidecar = self.validator(effective, self.run_id)
        self._assert_validation_sidecar(sidecar, effective)
        sidecar_path = self.run_dir / "05_validation_sidecar.json"
        _write_json(sidecar_path, sidecar)
        sidecar_sha = hashlib.sha256(sidecar_path.read_bytes()).hexdigest()
        _write_json(self.run_dir / "05_validation.json", {
            "stage": "validation",
            "status": "ok",
            "run_id": sidecar.get("run_id"),
            "sidecar_path": self._rel(sidecar_path),
            "sidecar_sha256": sidecar_sha,
            "n_strategies": len(sidecar.get("strategies") or []),
            "secondaries": sorted(normalize_ticker(s) for s in effective),
            "no_network": True,
            "no_validation_subprocess_in_dry_run": True,
        })

        # ----- Stage 6: join (run-id-bound k6 ranking; injectable seam) -----
        # The default joiner raises ValidationJoinError on any sidecar
        # SHA/run/row refusal; wrap it so the failure routes through the
        # existing _halt("execute", ...) envelope instead of escaping.
        ranking_path = self._resolve_k6_ranking_path(r4)
        try:
            v2 = self.joiner(ranking_path, sidecar_path, sidecar_sha)
        except _ValidationJoinError as exc:
            raise CrunchError("validation join failed: " + str(exc)) from exc
        if not isinstance(v2, dict):
            raise CrunchError("join did not return a v2 object")
        fresh_rows = v2.get("per_secondary")
        if not isinstance(fresh_rows, list) or not fresh_rows:
            raise CrunchError("join produced no built-only per_secondary rows")
        # Fail-closed normalization: the k6 ranking (and therefore the joined
        # fresh rows) may carry absolute path fields that promote's v2 validator
        # (run by combine's self-check) rejects. Normalize ONLY the fresh rows'
        # path fields to project-relative output/... BEFORE combine. The raw
        # on-disk k6 ranking artifact is NOT rewritten.
        fresh_rows, path_norm = self._normalize_publish_fresh_row_paths(fresh_rows)
        join_meta = {
            "stage": "join",
            "status": "ok",
            "ranking_artifact_path": self._rel(ranking_path),
            "expected_validation_sidecar_sha256": sidecar_sha,
            "built_row_count": len(fresh_rows),
            "secondaries": sorted(
                normalize_ticker(r.get("secondary")) for r in fresh_rows
                if isinstance(r, dict)),
            "path_normalization": path_norm,
            "no_public_fixture_write": True,
        }
        _write_json(self.run_dir / "06_join.json", join_meta)

        # ----- Stage 7: combine/proof -----
        if self.publish_fresh_ccc_records_file is None:
            reason = (
                "verified fresh CCC records are required for "
                "combine_and_assemble, but Blob upload/GET is disabled in "
                "publish dry-run; supply "
                "--publish-dry-run-fresh-ccc-records-file with already-verified "
                "records (get_verified=true) to exercise combine")
            _write_json(self.run_dir / "07_combine.json", {
                "stage": "combine",
                "status": "blocked",
                "combine_called": False,
                "reason": reason,
                "no_blob_upload": True,
                "no_blob_get": True,
                "no_synthesized_get_verified": True,
            })
            # Write the gate report (blocked) before halting, when practical.
            _write_json(self.run_dir / "08_publish_gate.json",
                        self._publish_gate_report(status="blocked", reason=reason,
                                                  candidate=None,
                                                  join_meta=join_meta))
            raise CrunchError("publish dry-run blocked at Stage 7: " + reason)

        fresh_ccc = self._load_fresh_ccc_records(
            self.publish_fresh_ccc_records_file)
        pub_dir = self.run_dir / "publish_candidate"
        try:
            summary = self.combiner(
                prior_fixture_path=self._prior_fixture_path(),
                prior_promotion_manifest_path=self._prior_promotion_manifest_path(),
                prior_ccc_verification_manifest_path=(
                    self.publish_prior_ccc_verification_manifest),
                fresh_rows=fresh_rows,
                fresh_validation_sidecar=sidecar,
                fresh_ccc_records=fresh_ccc,
                assembly_run_id=self.run_id,
                assembled_at_utc=self.now.strftime("%Y-%m-%dT%H:%M:%SZ"),
                output_dir=pub_dir,
                excluded_tickers=tuple(sorted(self._excl_set)),
                prior_validation_sidecar_path=(
                    self.publish_prior_validation_sidecar),
                project_root=self.project_root,
                run_self_check=True,
                reverify_carried_ccc=False,
            )
        except CombineError as exc:
            raise CrunchError(
                "combine/proof assembly failed: " + str(exc)) from exc
        if not isinstance(summary, dict):
            raise CrunchError("combine_and_assemble did not return a summary")
        _write_json(self.run_dir / "07_combine.json", {
            "stage": "combine",
            "status": "ok",
            "combine_called": True,
            "output_dir": self._rel(pub_dir),
            "summary": summary,
        })

        # ----- Stage 8: publish-gate dry-run -----
        gate = self._publish_gate_report(status="ok", reason=None,
                                         candidate=summary, join_meta=join_meta)
        _write_json(self.run_dir / "08_publish_gate.json", gate)
        return gate

    # --- default publish-tail seams (operator-only; never run by tests) ---
    def _default_validator(self, secondaries: list[str], run_id: str) -> dict:
        """Real Stage-5 validation over the rebuilt secondaries -> sidecar
        contract. Operator execution only; never invoked during the dry-run
        task or in tests (which inject a stub). Follows the adapter's real
        contract: build_adapter_inputs -> K6MtfValidationAdapter(secondaries,
        secondary_inputs) -> run_validation. Fail-closed on a non-mapping
        result/contract."""
        from utils.k6_mtf_validation.adapter import (  # noqa: PLC0415
            build_adapter_inputs, K6MtfValidationAdapter, run_validation)
        secs = list(secondaries)
        inputs = build_adapter_inputs(
            secs, stackbuilder_root=self.stackbuilder_root.as_posix())
        adapter = K6MtfValidationAdapter(secondaries=secs,
                                         secondary_inputs=inputs)
        result = run_validation(
            adapter, run_id=run_id,
            output_dir=self.run_dir / "publish_candidate" / "validation")
        if not isinstance(result, dict):
            raise CrunchError(
                "run_validation did not return a mapping result")
        contract = result.get("contract")
        if not isinstance(contract, dict):
            raise CrunchError(
                "run_validation result has no mapping 'contract'")
        return contract

    def _default_joiner(self, ranking_path: Path, sidecar_path: Path,
                        sidecar_sha: str) -> dict:
        from utils.react_publish.k6_mtf_validation_join import (  # noqa: PLC0415
            load_and_build_k6_mtf_ranking_v2)
        return load_and_build_k6_mtf_ranking_v2(
            ranking_path, sidecar_path,
            expected_validation_sidecar_sha256=sidecar_sha)

    def _default_combiner(self, **kwargs) -> dict:
        from crunch_combine_proof import combine_and_assemble  # noqa: PLC0415
        return combine_and_assemble(**kwargs)

    # --- Stage 9 publish tail (operator-launched; real publication) --------
    def _default_stage9_runner(self, inputs) -> dict:
        from stage9_publish import run_stage9_publish  # noqa: PLC0415
        return run_stage9_publish(inputs)

    def _git_toplevel(self) -> Path:
        """Git repository top-level (the publication + report-pair files live
        under it). Falls back to the project_root parent if git is unavailable."""
        try:
            out = subprocess.run(
                ["git", "-C", str(self.project_root), "rev-parse",
                 "--show-toplevel"],
                capture_output=True, text=True)
            top = (out.stdout or "").strip()
            if out.returncode == 0 and top:
                return Path(top)
        except Exception:  # noqa: BLE001
            pass
        return self.project_root.parent

    # NOTE: the historical _stage9_inputs(...) builder was unified into the
    # module-level run_rerank_publish seam (single source for Stage9PublishInputs
    # construction, with dry_run threaded). _run_stage9_publish_tail calls that
    # seam with dry_run=False, preserving Build-and-Rank behavior exactly.

    def _stage9_preflight(self) -> None:
        """Run the Stage 9 preflight BEFORE Stage 1 (no lock held through the
        build). Builds a checks-only inputs with placeholder fresh artifacts."""
        from stage9_publish import (  # noqa: PLC0415
            verify_publish_preflight, Stage9PublishInputs)
        fixtures_dir = self.project_root / "frontend" / "public" / "fixtures"
        pin = Stage9PublishInputs(
            repo_root=self._git_toplevel(),
            run_dir=self.run_dir,
            run_id=self.run_id,
            fresh_secondaries=(),
            fresh_rows=(),
            fresh_validation_sidecar={},
            k6_ranking_path=self.run_dir,
            prior_fixture_path=self._prior_fixture_path(),
            prior_promotion_manifest_path=self._prior_promotion_manifest_path(),
            prior_validation_sidecar_path=self.publish_prior_validation_sidecar,
            prior_ccc_verification_manifest_path=(
                self.publish_prior_ccc_verification_manifest),
            candidate_dir=self.run_dir / "publish_candidate",
            fresh_ccc_records_path=self.run_dir / "fresh_ccc_records.json",
            public_fixture_dest=fixtures_dir / "k6_mtf_ranking.json",
            public_manifest_dest=(
                fixtures_dir / "k6_mtf_ranking.promotion_manifest.json"),
            md_library_shared_dir=self.project_root / "md_library" / "shared",
            project_root=self.project_root,
            operator_approved=bool(self.operator_approved_publish),
            dry_run=False,
        )
        verify_publish_preflight(pin, acquire_lock=False)

    def _run_stage9_publish_tail(self, r4: dict, rebuild: list[str]) -> dict:
        """Stages 5-6 (validation + join + normalize) then the real Stage 9
        publication. Thin caller of the shared module-level run_rerank_publish
        seam with dry_run=False (Build-and-Rank's historical behavior). The
        ranking path is resolved from this run's Stage-4 result before the
        seam, which takes an explicit k6_ranking_path."""
        return run_rerank_publish(
            survivors=list(rebuild),
            k6_ranking_path=self._resolve_k6_ranking_path(r4),
            prior_fixture_path=self._prior_fixture_path(),
            prior_promotion_manifest_path=self._prior_promotion_manifest_path(),
            prior_validation_sidecar_path=self.publish_prior_validation_sidecar,
            prior_ccc_verification_manifest_path=(
                self.publish_prior_ccc_verification_manifest),
            run_dir=self.run_dir,
            run_id=self.run_id,
            target_as_of=self.target_as_of,
            dry_run=False,
            operator_approved=bool(self.operator_approved_publish),
            seams=RerankPublishSeams(
                validator=self.validator,
                joiner=self.joiner,
                stage9_runner=self._stage9_runner,
                project_root=self.project_root,
                repo_root=self._git_toplevel(),
                excluded_tickers=tuple(sorted(self._excl_set)),
            ),
        )

    # --- artifact path helpers + quarantine -------------------------------
    def _sec_dir_name(self, secondary: str) -> str:
        return secondary  # secondaries are literal dir names (incl carets)

    def _impactsearch_artifacts(self, rebuild: list[str]) -> list[Path]:
        out: list[Path] = []
        for sec in rebuild:
            out.extend(sorted(self.impactsearch_root.glob(f"{sec}_*.xlsx")))
            out.extend(sorted(self.impactsearch_root.glob(f"{sec}_*.json")))
        return out

    def _quarantine_impactsearch(self, rebuild: list[str]) -> dict:
        moved: dict[str, list[str]] = {}
        for sec in rebuild:
            items = sorted(self.impactsearch_root.glob(f"{sec}_*.xlsx"))
            items += sorted(self.impactsearch_root.glob(f"{sec}_*.json"))
            dest = self.run_dir / "quarantine" / "impactsearch" / sec
            m = quarantine_paths(items, dest)
            if m:
                moved[sec] = m
        return moved

    def _stackbuilder_artifacts(self, rebuild: list[str]) -> list[Path]:
        out: list[Path] = []
        for sec in rebuild:
            d = self.stackbuilder_root / self._sec_dir_name(sec)
            out.append(d / SELECTED_BUILD)
            out.append(d / SELECTED_BUILD_PINNED)
            out.extend(sorted(d.rglob(COMBO_FILENAME)))
        return out

    def _quarantine_stackbuilder(self, rebuild: list[str]) -> dict:
        moved: dict[str, list[str]] = {}
        for sec in rebuild:
            d = self.stackbuilder_root / self._sec_dir_name(sec)
            if d.exists():
                dest = self.run_dir / "quarantine" / "stackbuilder"
                m = quarantine_paths([d], dest)
                if m:
                    moved[sec] = m
        return moved

    def _assert_no_blocking_pins(self, rebuild: list[str]) -> None:
        pinned = []
        for sec in rebuild:
            p = self.stackbuilder_root / self._sec_dir_name(sec) / SELECTED_BUILD_PINNED
            if p.is_file():
                pinned.append(sec)
            sb = self.stackbuilder_root / self._sec_dir_name(sec) / SELECTED_BUILD
            if sb.is_file():
                try:
                    data = json.loads(sb.read_text(encoding="utf-8"))
                    if data.get("operator_pinned") is True:
                        pinned.append(sec)
                except (OSError, json.JSONDecodeError):
                    pass
        if pinned:
            raise CrunchError(
                "pinned selected_build blocks rebuild for "
                + json.dumps(sorted(set(pinned)), sort_keys=True)
                + "; --unpin not used without explicit operator approval")

    def _abspath(self, value: str) -> Path:
        p = Path(value)
        return p if p.is_absolute() else (self.project_root / p)

    def _resolve_k6_ranking_path(self, r4: dict) -> Path:
        """Resolve the EXACT current-run k6 ranking artifact, fail-closed and
        BOUND to the orchestrator run id. Because the orchestrator pins
        --driver-run-id <self.run_id>, the only acceptable artifact is
        ``<k6_output_root>/<self.run_id>/k6_mtf_ranking.json``. Every path or
        id the result returns must point exactly there; anything else (a
        stale/different run, even if its file exists) is a hard STOP. There is
        NO 'under the output root' acceptance and NO latest-run fallback."""
        stagef = r4.get("stageF") or {}
        expected_run_dir = self.k6_output_root / self.run_id
        expected_ranking = expected_run_dir / "k6_mtf_ranking.json"
        exp_dir = expected_run_dir.resolve()
        exp_rank = expected_ranking.resolve()

        # 1) driver_run_id, if returned, must equal the pinned run id.
        drid = r4.get("driver_run_id")
        if drid is not None and str(drid) != self.run_id:
            raise CrunchError(
                f"k6_recook driver_run_id {str(drid)!r} does not match the "
                f"orchestrator run id {self.run_id!r}")

        # 2) stageF.ranking_artifact_path, if present, must be the exact file.
        rp = stagef.get("ranking_artifact_path")
        if rp and self._abspath(str(rp)).resolve() != exp_rank:
            raise CrunchError(
                "k6_recook stageF.ranking_artifact_path points to a different "
                f"run than expected: {str(rp)!r} != {expected_ranking.as_posix()}")

        # 3) any returned run dir must be the exact expected run dir.
        for key, val in (("output_run_dir", r4.get("output_run_dir")),
                         ("run_dir", r4.get("run_dir")),
                         ("stageF.run_dir", stagef.get("run_dir"))):
            if val and self._abspath(str(val)).resolve() != exp_dir:
                raise CrunchError(
                    f"k6_recook {key} points to a different run than expected: "
                    f"{str(val)!r} != {expected_run_dir.as_posix()}")

        # 4) Fall back only to the expected (run-id-bound) artifact; require it.
        if not expected_ranking.is_file():
            raise CrunchError(
                "k6_recook ok but the current-run ranking artifact is missing "
                f"at {expected_ranking.as_posix()}")
        return expected_ranking

    def _k6_artifacts(self, r4: dict) -> list[Path]:
        ranking = self._resolve_k6_ranking_path(r4)
        out = [ranking]
        out.extend(sorted(ranking.parent.glob("*/k6_mtf_history.json")))
        return out


# ---------------------------------------------------------------------------
# Re-rank publish seam (module-level; importable without side effects)
#
# This is the SHARED real-publish tail extracted verbatim from
# CrunchOrchestrator._run_stage9_publish_tail so a thin re-rank driver can call
# it directly. The chain is: validator -> sidecar coverage assertion -> joiner
# -> fresh-row path normalization -> Stage9PublishInputs -> run_stage9_publish.
# dry_run is threaded into Stage9PublishInputs (the orchestrator's previous tail
# hardcoded dry_run=False; that is preserved at its call site below).
#
# The pure helpers (_normalize_one_output_path, _normalize_fresh_row_paths,
# _assert_validation_sidecar_covers) are the single source of truth: the
# matching CrunchOrchestrator methods now delegate to them, so the dry-run tail
# and the real-publish tail share identical behavior. The engine seams
# (validator/joiner/stage9_runner) are injected via RerankPublishSeams; the
# heavy modules (combine, stage9_publish, the join) are imported lazily inside
# the function so importing this module stays side-effect-free.
# ---------------------------------------------------------------------------


_RERANK_FRESH_ROW_PATH_FIELDS = (
    "history_artifact_path",
    "k6_stack.selected_build_path",
    "k6_stack.selected_run_dir",
    "k6_stack.combo_k6_path",
)


def _normalize_one_output_path(sec: str, field: str, value: Any, *,
                               project_root: Path) -> str:
    """Return a clean project-relative POSIX path under ``output/`` for one
    fresh-row path field, or raise CrunchError. Fail-closed: missing/empty/
    non-string, absolute-outside-output, relative-not-under-output, and
    traversal all hard-stop. Accepts absolute paths under
    ``<project_root>/output`` (any drive casing / slash style) and already-
    relative ``output/...`` paths (slashes normalized to POSIX)."""
    if not isinstance(value, str) or not value.strip():
        raise CrunchError(
            f"publish normalize: secondary {sec!r} field {field} is "
            f"missing/empty/non-string: {value!r}")
    s = value.replace("\\", "/").strip()
    is_abs = bool(re.match(r"^[A-Za-z]:", s)) or s.startswith("/") \
        or Path(s).is_absolute()
    if is_abs:
        out_root = (Path(project_root) / "output")
        try:
            rel = Path(value).resolve().relative_to(out_root.resolve())
        except (ValueError, OSError) as exc:
            raise CrunchError(
                f"publish normalize: secondary {sec!r} field {field} is an "
                "absolute path outside <project_root>/output: "
                f"{value!r}") from exc
        norm = "output/" + rel.as_posix()
    else:
        if not s.startswith("output/"):
            raise CrunchError(
                f"publish normalize: secondary {sec!r} field {field} is a "
                f"relative path not under output/: {value!r}")
        norm = s
    parts = norm.split("/")
    if parts[0] != "output" or ".." in parts:
        raise CrunchError(
            f"publish normalize: secondary {sec!r} field {field} did not "
            f"resolve to a clean output/ path (traversal/outside): {value!r}")
    return norm


def _normalize_fresh_row_paths(rows: list, *, project_root: Path) -> tuple:
    """Return (normalized_fresh_rows, summary). Normalizes ONLY the four path
    fields on each fresh row to project-relative POSIX ``output/...`` form;
    every other field (validation, metrics, CCC) is preserved exactly. Returns
    copies (the joiner return is not mutated in place). Carried rows are NOT
    touched here -- combine handles them from the prior fixture. Fail-closed via
    _normalize_one_output_path."""
    out_rows: list = []
    fields_normalized = 0
    rows_changed = 0
    for r in rows:
        if not isinstance(r, dict):
            raise CrunchError("publish normalize: fresh row is not an object")
        sec = normalize_ticker(r.get("secondary"))
        ks = r.get("k6_stack")
        if not isinstance(ks, dict):
            raise CrunchError(
                f"publish normalize: secondary {sec!r} k6_stack is missing "
                "or not a mapping")
        new_row = dict(r)
        new_ks = dict(ks)
        changed = False
        orig = r.get("history_artifact_path")
        norm = _normalize_one_output_path(
            sec, "history_artifact_path", orig, project_root=project_root)
        if norm != orig:
            changed = True
            fields_normalized += 1
        new_row["history_artifact_path"] = norm
        for f in ("selected_build_path", "selected_run_dir", "combo_k6_path"):
            o = ks.get(f)
            n = _normalize_one_output_path(
                sec, "k6_stack." + f, o, project_root=project_root)
            if n != o:
                changed = True
                fields_normalized += 1
            new_ks[f] = n
        new_row["k6_stack"] = new_ks
        if changed:
            rows_changed += 1
        out_rows.append(new_row)
    summary = {
        "fresh_rows_total": len(out_rows),
        "fresh_rows_normalized_count": rows_changed,
        "normalized_path_fields_count": fields_normalized,
        "normalized_path_fields": list(_RERANK_FRESH_ROW_PATH_FIELDS),
    }
    return out_rows, summary


def _assert_validation_sidecar_covers(sidecar: Any, effective: list[str], *,
                                      run_id: str) -> None:
    """Fail closed unless the Stage-5 sidecar covers EXACTLY the built
    secondaries with strict k6_mtf:<SEC> strategy ids and binds to this run
    id."""
    if not isinstance(sidecar, dict):
        raise CrunchError("validation sidecar is not a JSON object")
    if sidecar.get("validation_status") != "valid":
        raise CrunchError(
            "validation sidecar validation_status != 'valid': "
            f"{sidecar.get('validation_status')!r}")
    if sidecar.get("run_id") != run_id:
        raise CrunchError(
            "validation sidecar run_id "
            f"{sidecar.get('run_id')!r} != current run id {run_id!r}")
    strategies = sidecar.get("strategies")
    if not isinstance(strategies, list) or not strategies:
        raise CrunchError("validation sidecar 'strategies' must be a "
                          "non-empty list")
    seen: set[str] = set()
    for s in strategies:
        if not isinstance(s, dict):
            raise CrunchError("validation sidecar strategy is not an object")
        sid = s.get("strategy_id")
        if not isinstance(sid, str) or not sid.startswith("k6_mtf:"):
            raise CrunchError(
                f"validation sidecar strategy_id malformed: {sid!r}")
        sec = normalize_ticker(sid[len("k6_mtf:"):])
        if not sec or sid != f"k6_mtf:{sec}":
            raise CrunchError(
                f"validation sidecar strategy_id secondary mismatch: {sid!r}")
        if sec in seen:
            raise CrunchError(
                f"validation sidecar duplicate strategy secondary: {sec!r}")
        seen.add(sec)
    want = {normalize_ticker(s) for s in effective}
    if seen != want:
        missing = sorted(want - seen)
        extra = sorted(seen - want)
        raise CrunchError(
            "validation sidecar does not cover exactly the built "
            f"secondaries; missing {missing!r}, extra {extra!r}")


@dataclass
class RerankPublishSeams:
    """Injected engine seams + run-invariant config for run_rerank_publish.

    validator/joiner/stage9_runner mirror the orchestrator's publish-tail seams
    (CrunchOrchestrator.validator / .joiner / ._stage9_runner). project_root,
    repo_root, and excluded_tickers supply the Stage9PublishInputs fields the
    orchestrator otherwise reads from self."""
    validator: Callable[[list, str], dict]
    joiner: Callable[[Path, Path, str], dict]
    stage9_runner: Callable[[Any], dict]
    project_root: Path
    repo_root: Path
    excluded_tickers: Iterable[str] = ()


def run_rerank_publish(*, survivors: Iterable[str], k6_ranking_path: Path,
                       prior_fixture_path: Path,
                       prior_promotion_manifest_path: Path,
                       prior_validation_sidecar_path: Any,
                       prior_ccc_verification_manifest_path: Any,
                       run_dir: Path, run_id: str, target_as_of: Any,
                       dry_run: bool, operator_approved: bool,
                       seams: RerankPublishSeams) -> dict:
    """Run the shared real-publish tail over a surviving fresh set and return
    the Stage 9 summary dict. Behavior is identical to the orchestrator's
    historical _run_stage9_publish_tail except that ``dry_run`` is threaded into
    Stage9PublishInputs (previously hardcoded False).

    ``target_as_of`` is accepted for caller-symmetry with the re-rank driver and
    recorded by upstream stages; the publish chain itself derives all dates from
    the joined artifacts, so it is not consumed here. No network/Blob/promote/
    git happens in this function -- those live inside the injected
    ``stage9_runner`` (the real one is stage9_publish.run_stage9_publish, which
    is itself fail-closed and honors dry_run)."""
    from stage9_publish import Stage9PublishInputs  # noqa: PLC0415
    from utils.react_publish.k6_mtf_validation_join import (  # noqa: PLC0415
        ValidationJoinError as _ValidationJoinError)

    run_dir = Path(run_dir)
    effective = list(survivors)

    # ----- Stage 5: validation (injected seam) + strict coverage assertion ---
    sidecar = seams.validator(effective, run_id)
    _assert_validation_sidecar_covers(sidecar, effective, run_id=run_id)
    sidecar_path = run_dir / "05_validation_sidecar.json"
    _write_json(sidecar_path, sidecar)
    sidecar_sha = hashlib.sha256(sidecar_path.read_bytes()).hexdigest()

    # ----- Stage 6: join (run-id-bound k6 ranking) + path normalization ------
    ranking_path = Path(k6_ranking_path)
    try:
        v2 = seams.joiner(ranking_path, sidecar_path, sidecar_sha)
    except _ValidationJoinError as exc:
        raise CrunchError("validation join failed: " + str(exc)) from exc
    if not isinstance(v2, dict):
        raise CrunchError("join did not return a v2 object")
    fresh_rows = v2.get("per_secondary")
    if not isinstance(fresh_rows, list) or not fresh_rows:
        raise CrunchError("join produced no built-only per_secondary rows")
    fresh_rows, _ = _normalize_fresh_row_paths(
        fresh_rows, project_root=seams.project_root)

    # ----- Stage 9: real publication (combine happens inside the runner) -----
    fixtures_dir = Path(seams.project_root) / "frontend" / "public" / "fixtures"
    inputs = Stage9PublishInputs(
        repo_root=seams.repo_root,
        run_dir=run_dir,
        run_id=run_id,
        fresh_secondaries=tuple(sorted(normalize_ticker(s)
                                       for s in effective)),
        fresh_rows=fresh_rows,
        fresh_validation_sidecar=sidecar,
        k6_ranking_path=ranking_path,
        prior_fixture_path=prior_fixture_path,
        prior_promotion_manifest_path=prior_promotion_manifest_path,
        prior_validation_sidecar_path=prior_validation_sidecar_path,
        prior_ccc_verification_manifest_path=(
            prior_ccc_verification_manifest_path),
        candidate_dir=run_dir / "publish_candidate",
        fresh_ccc_records_path=run_dir / "fresh_ccc_records.json",
        public_fixture_dest=fixtures_dir / "k6_mtf_ranking.json",
        public_manifest_dest=(
            fixtures_dir / "k6_mtf_ranking.promotion_manifest.json"),
        md_library_shared_dir=(
            Path(seams.project_root) / "md_library" / "shared"),
        project_root=seams.project_root,
        excluded_tickers=tuple(sorted(seams.excluded_tickers)),
        operator_approved=bool(operator_approved),
        dry_run=bool(dry_run),
    )
    result = seams.stage9_runner(inputs)
    if not isinstance(result, dict):
        raise CrunchError("stage 9 runner did not return a summary")
    _write_json(run_dir / "09_stage9_publish.json", result)
    return result


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="crunch_rebuild_orchestrator",
        description=(
            "Run-once crunch rebuild orchestrator. Dry-run by default "
            "(Stage 0 preflight + plan only; no engines, no network). "
            "--execute (operator-run only) chains onepass -> impactsearch -> "
            "stackbuilder -> k6_recook with exclusion boundary checks and "
            "quarantine, and STOPS before any publication."
        ),
    )
    p.add_argument("--blocked-tickers-file", default=DEFAULT_BLOCKED_FILE)
    p.add_argument("--rebuild-secondaries-file", default=DEFAULT_REBUILD_FILE)
    p.add_argument("--run-dir", default=None)
    p.add_argument("--target-as-of", default=None)
    p.add_argument("--duration-budget-minutes", type=int, default=None)
    p.add_argument("--operator-budget-label", default=None)
    p.add_argument("--allow-network-fetch", action="store_true")
    p.add_argument("--execute", action="store_true")
    p.add_argument("--reclaim-stale-lock", action="store_true")
    p.add_argument("--reuse-onepass-run-dir", default=None,
                   help="Reuse a prior crunch run dir's validated OnePass "
                        "evidence instead of re-running Stage 1 (opt-in; "
                        "fail-closed; absolute or project-relative path).")
    p.add_argument("--reuse-onepass-max-age-hours", type=int,
                   default=DEFAULT_REUSE_ONEPASS_MAX_AGE_HOURS,
                   help="Freshness window (hours) for OnePass reuse "
                        f"(default {DEFAULT_REUSE_ONEPASS_MAX_AGE_HOURS}).")
    p.add_argument("--stackbuilder-root", default=DEFAULT_STACKBUILDER_ROOT)
    p.add_argument("--impactsearch-root", default=DEFAULT_IMPACTSEARCH_ROOT)
    p.add_argument("--onepass-root", default=DEFAULT_ONEPASS_ROOT)
    p.add_argument("--k6-output-root", default=DEFAULT_K6_OUTPUT_ROOT)
    p.add_argument("--master-tickers-file", default=DEFAULT_MASTER_TICKERS)
    p.add_argument("--project-root", default=None)
    # Publish-dry-run tail (Stages 5-8). Publication boundary stays CLOSED:
    # no Blob, no promote write, no public fixture write, no commit/push/deploy.
    p.add_argument("--publish-dry-run", action="store_true",
                   help="After a clean Stage 4, run the gated publish-dry-run "
                        "tail (validation -> join -> combine/proof -> publish "
                        "gate). Writes candidate artifacts + a would-be publish "
                        "plan under the run dir and STOPS before any "
                        "publication. No Blob/promote/fixture/commit/push.")
    p.add_argument("--publish-dry-run-fresh-ccc-records-file", default=None,
                   help="Read-only JSON file of ALREADY-verified fresh CCC "
                        "records (get_verified=true) used only to exercise "
                        "combine/proof without Blob upload/GET. If absent, the "
                        "tail fails closed at Stage 7.")
    p.add_argument("--publish-prior-fixture", default=None)
    p.add_argument("--publish-prior-promotion-manifest", default=None)
    p.add_argument("--publish-prior-validation-sidecar", default=None)
    p.add_argument("--publish-prior-ccc-verification-manifest", default=None)
    p.add_argument("--publish", action="store_true",
                   help="OPERATOR-LAUNCHED real publish: after a clean Stage 4, "
                        "run the Stage 9 publish tail (same-run CCC upload -> "
                        "combine -> promote dry-run -> promote write -> commit "
                        "-> push -> live verify), fail-closed. Requires "
                        "--execute and --operator-approved-publish. Crosses the "
                        "publication boundary; do not run inside Claude Code.")
    p.add_argument("--operator-approved-publish", action="store_true",
                   help="Explicit operator approval gate REQUIRED for --publish "
                        "to perform the real Blob upload, public fixture write, "
                        "commit, and push.")
    return p


def _resolve(project_root: Path, opt: str, default_rel: str) -> Path:
    q = Path(opt) if opt else Path(default_rel)
    return q if q.is_absolute() else (project_root / q)


def main(argv: Iterable[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(list(argv) if argv is not None else None)
    if args.publish and args.publish_dry_run:
        parser.error("--publish and --publish-dry-run are mutually exclusive")
    if args.publish and not args.execute:
        parser.error("--publish requires --execute")
    if args.publish and not args.operator_approved_publish:
        parser.error("--publish requires --operator-approved-publish")
    project_root = (Path(args.project_root).resolve() if args.project_root
                    else Path(__file__).resolve().parent)
    now = _utcnow()
    run_dir = (Path(args.run_dir) if args.run_dir else
               project_root / DEFAULT_RUN_BASE / _run_id(now))
    if not run_dir.is_absolute():
        run_dir = project_root / run_dir

    reuse_dir = (_resolve(project_root, args.reuse_onepass_run_dir, "")
                 if args.reuse_onepass_run_dir else None)

    def _opt_path(opt):
        return _resolve(project_root, opt, "") if opt else None

    orch = CrunchOrchestrator(
        project_root=project_root,
        run_dir=run_dir,
        blocked_file=_resolve(project_root, args.blocked_tickers_file,
                              DEFAULT_BLOCKED_FILE),
        rebuild_file=_resolve(project_root, args.rebuild_secondaries_file,
                              DEFAULT_REBUILD_FILE),
        stackbuilder_root=_resolve(project_root, args.stackbuilder_root,
                                   DEFAULT_STACKBUILDER_ROOT),
        impactsearch_root=_resolve(project_root, args.impactsearch_root,
                                   DEFAULT_IMPACTSEARCH_ROOT),
        onepass_root=_resolve(project_root, args.onepass_root,
                              DEFAULT_ONEPASS_ROOT),
        k6_output_root=_resolve(project_root, args.k6_output_root,
                                DEFAULT_K6_OUTPUT_ROOT),
        master_tickers_file=_resolve(project_root, args.master_tickers_file,
                                     DEFAULT_MASTER_TICKERS),
        target_as_of=args.target_as_of,
        duration_budget_minutes=args.duration_budget_minutes,
        operator_budget_label=args.operator_budget_label,
        allow_network_fetch=args.allow_network_fetch,
        execute=args.execute,
        reclaim_stale_lock=args.reclaim_stale_lock,
        now=now,
        reuse_onepass_run_dir=reuse_dir,
        reuse_onepass_max_age_hours=args.reuse_onepass_max_age_hours,
        publish_dry_run=args.publish_dry_run,
        publish_fresh_ccc_records_file=_opt_path(
            args.publish_dry_run_fresh_ccc_records_file),
        publish_prior_fixture=_opt_path(args.publish_prior_fixture),
        publish_prior_promotion_manifest=_opt_path(
            args.publish_prior_promotion_manifest),
        publish_prior_validation_sidecar=_opt_path(
            args.publish_prior_validation_sidecar),
        publish_prior_ccc_verification_manifest=_opt_path(
            args.publish_prior_ccc_verification_manifest),
        publish=args.publish,
        operator_approved_publish=args.operator_approved_publish,
    )
    env = orch.run()
    print(json.dumps({
        "status": env.get("status"),
        "halted_at": env.get("halted_at"),
        "run_dir": run_dir.as_posix(),
        "mode": env.get("mode"),
    }, indent=2, sort_keys=True))
    return 0 if env.get("status") in ("dry_run_planned",
                                      "completed_no_publish",
                                      "completed_publish_dry_run",
                                      "completed_publish") else 1


if __name__ == "__main__":
    sys.exit(main())
